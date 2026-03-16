import openpyxl

wb = openpyxl.load_workbook('CALENDARIO 2026.xlsx', data_only=True)
print('Hojas:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} === (filas: {ws.max_row}, cols: {ws.max_column})')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=False):
        vals = [(c.value, c.coordinate) for c in row if c.value is not None]
        if vals:
            print(vals)
