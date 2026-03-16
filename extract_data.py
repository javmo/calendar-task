import openpyxl
import json
from datetime import datetime

wb = openpyxl.load_workbook('CALENDARIO 2026.xlsx', data_only=True)

# --- Extract VTOS (Vencimientos) ---
ws_vtos = wb['VTOS']
task_types_vtos = []
for col in range(1, ws_vtos.max_column + 1, 2):
    name = ws_vtos.cell(row=1, column=col).value
    if name:
        task_types_vtos.append({
            'name': name,
            'name_col': col,
            'date_col': col + 1
        })

vencimientos = {}
for tt in task_types_vtos:
    vencimientos[tt['name']] = {}
    for row in range(2, ws_vtos.max_row + 1):
        cat = ws_vtos.cell(row=row, column=tt['name_col']).value
        date_val = ws_vtos.cell(row=row, column=tt['date_col']).value
        if cat is not None and date_val is not None:
            cat_key = str(int(cat))
            if isinstance(date_val, datetime):
                vencimientos[tt['name']][cat_key] = date_val.strftime('%Y-%m-%d')

# --- Extract TAREAS ---
ws_tareas = wb['TAREAS']
tareas = []
for row in range(3, ws_tareas.max_row + 1):
    cliente = ws_tareas.cell(row=row, column=2).value
    tarea = ws_tareas.cell(row=row, column=3).value
    quien = ws_tareas.cell(row=row, column=4).value
    semana = ws_tareas.cell(row=row, column=5).value
    if cliente and tarea:
        # Clean non-breaking spaces
        cliente_clean = str(cliente).replace('\xa0', ' ').strip()
        tareas.append({
            'id': row - 2,
            'cliente': cliente_clean,
            'tarea': str(tarea).strip(),
            'responsable': str(quien).strip() if quien else 'PERSONA',
            'semana': str(semana).strip() if semana else '1ER SEMANA',
            'completado': False,
            'revisado': False,
            'enviado': False
        })

# --- Assign due dates based on VTOS ---
# Map week to approximate date range for March 2026
week_dates = {
    '1ER SEMANA': '2026-03-02',
    '2DA SEMANA': '2026-03-09',
    '3ER SEMANA': '2026-03-16',
    '4TA SEMANA': '2026-03-23'
}

# Build case-insensitive lookup for vencimientos
vencimientos_lower = {k.lower(): v for k, v in vencimientos.items()}

# For each task, pick a due date from VTOS if possible, otherwise use week start
for t in tareas:
    tarea_name = t['tarea']
    tarea_lower = tarea_name.lower()
    if tarea_lower in vencimientos_lower:
        # Get some due date from the vencimientos (use category based on task id modulo)
        vtos_data = vencimientos_lower[tarea_lower]
        cat_key = str(t['id'] % 10)
        if cat_key in vtos_data:
            t['vencimiento'] = vtos_data[cat_key]
        else:
            # Use first available date
            dates = list(vtos_data.values())
            if dates:
                t['vencimiento'] = dates[0]
            else:
                t['vencimiento'] = week_dates.get(t['semana'], '2026-03-01')
    else:
        # For tasks not in VTOS (like Tareas generales, Liquidacion de sueldos)
        t['vencimiento'] = week_dates.get(t['semana'], '2026-03-01')

# Build data structure
data = {
    'clientes': sorted(list(set(t['cliente'] for t in tareas))),
    'tiposTarea': sorted(list(set(t['tarea'] for t in tareas))),
    'responsables': sorted(list(set(t['responsable'] for t in tareas))),
    'vencimientos': vencimientos,
    'tareas': tareas
}

# Write JSON
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f'Exported {len(tareas)} tareas for {len(data["clientes"])} clientes')
print(f'Tipos de tarea: {data["tiposTarea"]}')
print(f'Responsables: {data["responsables"]}')
print(f'Vencimientos keys: {list(vencimientos.keys())}')

# Show sample
for t in tareas[:5]:
    print(t)
