"""
Script para cargar la tabla de vencimientos del Excel a MongoDB.
Lee la hoja VTOS del Excel y crea documentos en la colección 'vencimientos'.
También recalcula los vencimientos de las tareas existentes.
"""
import os
from datetime import datetime, timezone
from pymongo import MongoClient
from dotenv import load_dotenv
import openpyxl

load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = os.getenv("DB_NAME", "calendario")

# Tipos de tarea que tienen vencimiento por dígito CUIT (orden de columnas en VTOS)
# Pares en el Excel: col_digito, col_fecha
TASK_TYPES_VTOS = [
    ("Casas Particulares", 0, 1),       # cols A,B
    # col C,D = "4° Anticipos" (skip)
    ("VEP Monotributo", 4, 5),          # cols E,F
    ("VEP Autonomos", 6, 7),            # cols G,H
    ("IIBB CM", 8, 9),                  # cols I,J
    ("IIBB ARBA", 10, 11),              # cols K,L
    ("IIBB AGIP", 12, 13),              # cols M,N
    ("IVA", 14, 15),                    # cols O,P
    ("Libro IVA Digital", 16, 17),      # cols Q,R
]


def read_vtos_from_excel(filepath):
    """Lee la hoja VTOS y retorna la tabla de vencimientos."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['VTOS']

    # Rows 3-12 (1-indexed) = digits 0-9
    # Header is row 2
    tabla = {}

    for task_name, digit_col, date_col in TASK_TYPES_VTOS:
        tabla[task_name] = {}
        for row_idx in range(3, 13):  # rows 3 to 12
            digit_cell = ws.cell(row=row_idx, column=digit_col + 1).value
            date_cell = ws.cell(row=row_idx, column=date_col + 1).value

            if digit_cell is not None:
                digit = str(int(float(digit_cell)))
                if isinstance(date_cell, datetime):
                    tabla[task_name][digit] = date_cell.strftime("%Y-%m-%d")
                elif date_cell and date_cell != "no hay":
                    tabla[task_name][digit] = str(date_cell)[:10]
                else:
                    tabla[task_name][digit] = None

    wb.close()
    return tabla


def get_periodo_from_excel(filepath):
    """Deduce el periodo (mes) del Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['VTOS']
    # Row 1, Col A has the reference date
    ref = ws.cell(row=1, column=1).value
    wb.close()
    if isinstance(ref, datetime):
        # The Excel is for the month AFTER the reference date
        month = ref.month + 1
        year = ref.year
        if month > 12:
            month = 1
            year += 1
        return f"{year}-{month:02d}"
    return "2026-03"


def main():
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]

    filepath = "CALEN 2026_03.xlsx"

    # Read vencimientos
    tabla = read_vtos_from_excel(filepath)
    periodo = get_periodo_from_excel(filepath)

    print(f"📅 Periodo: {periodo}")
    print(f"📋 Tipos de tarea con vencimientos: {len(tabla)}")
    for tipo, digitos in tabla.items():
        sample = digitos.get("0", "N/A")
        print(f"   {tipo}: dígito 0 → {sample}")

    # Create vencimientos collection
    now = datetime.now(timezone.utc).isoformat()

    # Upsert the vencimientos doc for this periodo
    db.vencimientos.update_one(
        {"periodo": periodo},
        {"$set": {
            "tabla": tabla,
            "updatedAt": now,
        },
        "$setOnInsert": {
            "createdAt": now,
        }},
        upsert=True,
    )
    print(f"\n✅ Vencimientos guardados para {periodo}")

    # Create indexes
    db.vencimientos.create_index("periodo", unique=True)

    # Now recalculate task vencimientos for this period
    # Get all tasks for this month
    mes_map = {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
    }
    mes_nombre = mes_map.get(periodo.split("-")[1], "")
    print(f"\n🔄 Recalculando vencimientos para tareas del mes: {mes_nombre}")

    # Get all clients with CUIT
    clientes = list(db.clientes.find({}, {"nombre": 1, "cuit": 1}))
    cuit_map = {}
    for c in clientes:
        if c.get("cuit"):
            cuit_str = str(c["cuit"]).replace(".", "").strip()
            if cuit_str:
                last_digit = cuit_str[-1]
                cuit_map[c["nombre"]] = last_digit

    print(f"   Clientes con CUIT: {len(cuit_map)}")

    # Update tasks
    updated_count = 0
    tasks = list(db.tasks.find({"mes": mes_nombre}))
    for task in tasks:
        cliente = task.get("cliente", "")
        tarea = task.get("tarea", "")

        if cliente not in cuit_map:
            continue
        if tarea not in tabla:
            continue  # "Tareas generales" etc. don't have vencimiento lookup

        digit = cuit_map[cliente]
        new_vto = tabla[tarea].get(digit)

        if new_vto and new_vto != task.get("vencimiento"):
            db.tasks.update_one(
                {"_id": task["_id"]},
                {"$set": {"vencimiento": new_vto}}
            )
            updated_count += 1

    print(f"✅ {updated_count} tareas actualizadas con nuevos vencimientos")

    # Also seed April through December with empty templates
    for m in range(4, 13):
        p = f"2026-{m:02d}"
        existing = db.vencimientos.find_one({"periodo": p})
        if not existing:
            empty_tabla = {}
            for tipo in tabla:
                empty_tabla[tipo] = {str(d): None for d in range(10)}
            db.vencimientos.update_one(
                {"periodo": p},
                {"$set": {"tabla": empty_tabla, "updatedAt": now},
                 "$setOnInsert": {"createdAt": now}},
                upsert=True,
            )
            print(f"   📝 Plantilla vacía creada para {p}")

    client.close()
    print("\n🎉 ¡Listo!")


if __name__ == "__main__":
    main()
