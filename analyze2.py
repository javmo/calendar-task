import openpyxl

wb = openpyxl.load_workbook('CALENDARIO 2026.xlsx', data_only=True)
ws = wb['TAREAS']
print(f'TAREAS: filas={ws.max_row}, cols={ws.max_column}')
for row in ws.iter_rows(min_row=60, max_row=ws.max_row, values_only=False):
    vals = [(c.value, c.coordinate) for c in row if c.value is not None]
    if vals:
        print(vals)

# Also get unique values for analysis
clients = set()
tasks = set()
owners = set()
weeks = set()
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    if row[1]: clients.add(str(row[1]).strip())
    if row[2]: tasks.add(str(row[2]).strip())
    if row[3]: owners.add(str(row[3]).strip())
    if row[4]: weeks.add(str(row[4]).strip())

print(f'\n--- RESUMEN ---')
print(f'Total clientes unicos: {len(clients)}')
print(f'Clientes: {sorted(clients)}')
print(f'\nTareas unicas: {sorted(tasks)}')
print(f'\nResponsables: {sorted(owners)}')
print(f'\nSemanas: {sorted(weeks)}')
